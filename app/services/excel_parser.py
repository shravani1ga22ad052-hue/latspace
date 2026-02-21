from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from app.registry import load_parameter_registry
from app.services.column_mapper import MappingResult, map_header
from app.services.head_detector import detect_header_row
from app.services.validator import build_validation_warnings
from app.services.value_parser import parse_value


def parse_workbook(binary: bytes) -> dict[str, Any]:
    registry = load_parameter_registry()
    workbook = load_workbook(io.BytesIO(binary), data_only=True)

    all_records: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue

        header_idx = detect_header_row(rows)
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_idx]]

        mappings: list[MappingResult] = []
        mapped_pairs: dict[tuple[str, str | None], int] = {}

        for idx, header in enumerate(headers):
            mapping = map_header(header, registry)
            mapping.column_index = idx
            mappings.append(mapping)

        data_row_count = 0
        for raw_row in rows[header_idx + 1 :]:
            if all(cell is None or str(cell).strip() == "" for cell in raw_row):
                continue

            for mapping in mappings:
                if not mapping.parameter:
                    continue
                raw_value = raw_row[mapping.column_index] if mapping.column_index < len(raw_row) else None
                record = {
                    "sheet": sheet.title,
                    "row_index": data_row_count,
                    "parameter": mapping.parameter,
                    "asset": mapping.asset,
                    "value": parse_value(raw_value),
                    "raw_value": raw_value,
                    "confidence": mapping.confidence,
                }
                all_records.append(record)
                pair = (mapping.parameter, mapping.asset)
                mapped_pairs[pair] = mapped_pairs.get(pair, 0) + 1
            data_row_count += 1

        unmapped = [m.original_header for m in mappings if m.original_header and not m.parameter]
        duplicate_columns = [
            {"parameter": parameter, "asset": asset, "count": count}
            for (parameter, asset), count in mapped_pairs.items()
            if data_row_count > 0 and count > data_row_count
        ]

        sheets.append(
            {
                "sheet": sheet.title,
                "header_row": header_idx + 1,
                "mapped_columns": [m.__dict__ for m in mappings if m.parameter],
                "unmapped_columns": unmapped,
                "duplicate_columns": duplicate_columns,
            }
        )

    return {
        "records": all_records,
        "sheets": sheets,
        "validation_warnings": build_validation_warnings(all_records),
        "record_count": len(all_records),
    }
