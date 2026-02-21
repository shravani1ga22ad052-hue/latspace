from pathlib import Path

from openpyxl import Workbook

OUT = Path("test_data")
OUT.mkdir(exist_ok=True)


wb = Workbook()
ws = wb.active
ws.title = "clean"
ws.append(["Date", "Coal Consumption", "Steam Generation", "Boiler Efficiency"])
ws.append(["2026-01-01", 100, 240, "88%"])
ws.append(["2026-01-02", 110, 245, "89%"])
wb.save(OUT / "clean_data.xlsx")


wb = Workbook()
ws = wb.active
ws.title = "messy"
ws.append(["Acme Pulp Unit Monthly Report"]) 
ws.append(["January 2026"]) 
ws.append([])
ws.append(["Dt", "COAL CONSMPTN", "Steam (Boiler 2)", "Effcy %", "Random Notes"])
ws.append(["01-Jan", "1,234.56", "250", "91%", "N/A"])
ws.append(["02-Jan", "1,200", "245", "89%", "Maintenance"])
wb.save(OUT / "messy_data.xlsx")


wb = Workbook()
ws = wb.active
ws.title = "assets"
ws.append(["Date", "Coal Consumption AFBC-1", "Coal Consumption AFBC-2", "Power TG1", "Power TG-2"])
ws.append(["2026-01-01", 300, 280, 40, 38])
ws.append(["2026-01-02", 305, 275, 42, 39])
wb.save(OUT / "multi_asset.xlsx")

print(f"Created sample files under {OUT.resolve()}")
